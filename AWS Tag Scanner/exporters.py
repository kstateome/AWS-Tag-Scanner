"""
Export functionality for AWS tag scan results.

This module handles all export formats and their specific formatting requirements:
- CSV with compliance annotations
- JSON
- Excel with multi-sheet workbooks
- Console output (detailed and summary)
"""

import csv
import json
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from security_utils import validate_output_filename, sanitize_csv_value
from report_builder import ReportBuilder


class ResourceExporter:
    """Handles exporting AWS resource data to various formats."""
    
    def __init__(self, report_builder: ReportBuilder):
        """
        Initialize exporter with a report builder for compliance checking.
        
        Args:
            report_builder: ReportBuilder instance for policy compliance
        """
        self.report_builder = report_builder
        self.logger = logging.getLogger(__name__)
    
    @property
    def tag_policy(self) -> Optional[Dict[str, Any]]:
        """Access tag policy from report builder"""
        return self.report_builder.tag_policy
    
    # ========================================================================
    # CSV EXPORT
    # ========================================================================
    
    def _generate_csv_filename(self, filename: Optional[str] = None) -> str:
        """Generate CSV filename with timestamp if not provided"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"aws_tags_scan_{timestamp}.csv"
        return filename
    
    def _collect_tag_columns(self, resources: List[Dict[str, Any]]) -> List[str]:
        """Collect all unique tag keys from resources and return sorted list"""
        all_tag_keys: set[str] = set()
        for resource in resources:
            all_tag_keys.update(resource['Tags'].keys())
        return sorted(all_tag_keys)
    
    def _build_csv_headers(self, tag_columns: List[str]) -> List[str]:
        """Build CSV header row"""
        headers = ['Service', 'ResourceType', 'ResourceId', 'ResourceName', 'State', 'TagCount']
        headers.extend(tag_columns)
        
        if self.tag_policy:
            headers.extend(['PolicyCompliant', 'ComplianceScore', 'MissingRequired', 
                          'MissingRecommended', 'UnneededTags'])
        return headers
    
    def _get_service_required_tags(self, resource: Dict[str, Any]) -> set[str]:
        """Get the set of tags required for a specific resource's service type"""
        if not self.tag_policy or 'Compliance' not in resource:
            return set()
        
        comp = resource['Compliance']
        service_type = comp.get('service_type', 'other')
        service_reqs = self.tag_policy.get('service_specific_requirements', {}).get(service_type, {})
        core_tags = self.tag_policy.get('core_tags', {}).get('tags', [])
        
        # Build set of tags that are valid for this service
        service_required_tags = {tag['key'] for tag in core_tags}
        if service_reqs:
            service_required_tags.update(service_reqs.get('required_tags', []))
            service_required_tags.update(service_reqs.get('recommended_tags', []))
        
        return service_required_tags
    
    def _format_csv_tag_value(self, tag_key: str, tag_value: str, 
                             service_required_tags: set[str]) -> str:
        """Format tag value for CSV based on policy requirements"""
        if not self.tag_policy or not service_required_tags:
            return tag_value
        
        normalized_key = self.report_builder.normalize_tag_key(tag_key)
        
        if normalized_key not in service_required_tags:
            # This tag is not needed for this service type
            if tag_value:
                return f"X - {tag_value}"
            else:
                return "X"
        
        # Tag is required/recommended - return as-is (blank if missing)
        return tag_value
    
    def _build_csv_row(self, resource: Dict[str, Any], tag_columns: List[str]) -> List[Any]:
        """Build a single CSV row for a resource"""
        row: List[Any] = [
            sanitize_csv_value(resource['Service']),
            sanitize_csv_value(resource['ResourceType']), 
            sanitize_csv_value(resource['ResourceId']),
            sanitize_csv_value(resource['ResourceName']),
            sanitize_csv_value(resource['State']),
            resource['TagCount']
        ]
        
        # Get service-specific required tags
        service_required_tags = self._get_service_required_tags(resource)
        
        # Add tag values
        for tag_key in tag_columns:
            tag_value = resource['Tags'].get(tag_key, '')
            formatted_value = self._format_csv_tag_value(tag_key, tag_value, service_required_tags)
            row.append(sanitize_csv_value(formatted_value))
        
        # Add compliance columns if available
        if self.tag_policy and 'Compliance' in resource:
            comp = resource['Compliance']
            row.append('YES' if comp['compliant'] else 'NO')
            row.append(f"{comp['compliance_score']:.1f}%")
            row.append('; '.join(comp['missing_required']) if comp['missing_required'] else '')
            row.append('; '.join(comp['missing_recommended']) if comp['missing_recommended'] else '')
            row.append('; '.join(comp['unneeded_tags']) if comp['unneeded_tags'] else '')
        
        return row
    
    def _print_csv_legend(self) -> None:
        """Print CSV legend explaining tag marking system"""
        self.logger.info("* Policy compliance columns included")
        self.logger.info("* Tag column legend:")
        self.logger.info("  - Blank cell = Tag IS required but missing from resource")
        self.logger.info("  - 'X' = Tag is NOT required for this service type")
        self.logger.info("  - Value shown = Tag exists and is appropriate")
        self.logger.info("* See 'MissingRequired' column for tags that need to be added")
    
    def export_to_csv(self, resources: List[Dict[str, Any]], filename: Optional[str] = None) -> str:
        """Export resources to CSV format"""
        # If a filename was supplied, validate it (no paths)
        if filename:
            try:
                clean = validate_output_filename(filename)
            except ValueError as e:
                raise
            # Ensure extension
            if not clean.lower().endswith('.csv'):
                clean = clean + '.csv'
            filename = clean
        else:
            filename = self._generate_csv_filename(filename)
        tag_columns = self._collect_tag_columns(resources)
        headers = self._build_csv_headers(tag_columns)
        
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(headers)
            
            # Write data rows
            for resource in resources:
                row = self._build_csv_row(resource, tag_columns)
                writer.writerow(row)
        
        self.logger.info(f"[CSV] Report exported to: {filename}")
        if self.tag_policy:
            self._print_csv_legend()
        
        return filename
    
    # ========================================================================
    # JSON EXPORT
    # ========================================================================
    
    def export_to_json(self, resources: List[Dict[str, Any]], filename: Optional[str] = None) -> str:
        """Export resources to JSON format"""
        if filename:
            try:
                clean = validate_output_filename(filename)
            except ValueError:
                raise
            if not clean.lower().endswith('.json'):
                clean = clean + '.json'
            filename = clean
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"aws_tags_scan_{timestamp}.json"

        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(resources, f, indent=2, default=str)

        self.logger.info(f"[JSON] Report exported to: {filename}")
        return filename

    # ========================================================================
    # EXCEL EXPORT
    # ========================================================================
    
    def _get_required_tags_for_service(self, service_type: str) -> List[str]:
        """Get list of required tags for a service type"""
        if not self.tag_policy:
            return []
        
        core_tags = self.tag_policy.get('core_tags', {}).get('tags', [])
        required = [tag['key'] for tag in core_tags]
        
        service_reqs = self.tag_policy.get('service_specific_requirements', {}).get(service_type, {})
        if service_reqs:
            required.extend(service_reqs.get('required_tags', []))
        
        return required
    
    def _get_recommended_tags_for_service(self, service_type: str) -> List[str]:
        """Get list of recommended tags for a service type"""
        if not self.tag_policy:
            return []
        
        service_reqs = self.tag_policy.get('service_specific_requirements', {}).get(service_type, {})
        if service_reqs:
            return service_reqs.get('recommended_tags', [])
        
        return []
    
    def _write_sheet_with_auto_width(self, ws, headers: List[str], data_rows: List[List[Any]]) -> None:
        """Write headers and data to sheet with auto-width columns"""
        # Write header
        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c_idx, value=h)
            cell.font = cell.font.copy(bold=True)
        
        # Write rows
        for r_idx, row in enumerate(data_rows, start=2):
            for c_idx, val in enumerate(row, start=1):
                cell_val = '' if val is None else str(val)
                ws.cell(row=r_idx, column=c_idx, value=cell_val)
        
        # Auto-width columns
        for i, _ in enumerate(headers, start=1):
            col = get_column_letter(i)
            max_len = 0
            for cell in ws[col]:
                if cell.value:
                    try:
                        l = len(str(cell.value))
                    except Exception:
                        l = 0
                    if l > max_len:
                        max_len = l
            ws.column_dimensions[col].width = min(max(10, max_len + 2), 60)
    
    def export_to_excel(self, resources: List[Dict[str, Any]], filename: Optional[str] = None) -> str:
        """
        Export resources into an Excel workbook with separate sheets per service.
        Each service sheet contains three sections:
        1. Required Tags - Shows required tags with 'MISSING' for absent ones
        2. Recommended Tags - Shows recommended tags with 'MISSING' for absent ones
        3. Current Tags - Shows all actual tags on resources
        """
        # validate/sanitize filename
        if filename:
            try:
                clean = validate_output_filename(filename)
            except ValueError:
                raise
            if not clean.lower().endswith('.xlsx'):
                clean = clean + '.xlsx'
            filename = clean
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"aws_tags_scan_{timestamp}.xlsx"

        # Group resources by service
        by_service: Dict[str, List[Dict[str, Any]]] = {}
        for r in resources:
            svc = r.get('Service', 'other')
            by_service.setdefault(svc, []).append(r)

        wb = Workbook()
        # remove default sheet
        default = wb.active
        wb.remove(default)

        # Create sheets per service with Required/Recommended/Current sections
        for service, items in sorted(by_service.items()):
            # Create three sub-sheets per service
            self._create_required_tags_sheet(wb, items, f"{service}-Required")
            self._create_recommended_tags_sheet(wb, items, f"{service}-Recommended")
            self._create_current_tags_sheet(wb, items, f"{service}-Current")

        wb.save(filename)
        self.logger.info(f"[XLSX] Report exported to: {filename}")
        self.logger.info(f"[XLSX]   Created {len(by_service)} services with 3 sheets each (Required/Recommended/Current)")
        for service in sorted(by_service.keys()):
            self.logger.info(f"[XLSX]   - {service}: {len(by_service[service])} resources")
        return filename
    
    def _create_required_tags_sheet(self, wb: Workbook, resources: List[Dict[str, Any]], sheet_name: str = "Required Tags") -> None:
        """Create sheet showing required tags with MISSING markers"""
        ws = wb.create_sheet(title=sheet_name[:31])
        
        # Collect all required tag columns across all resources
        required_tags_set = set()
        for resource in resources:
            if 'Compliance' in resource:
                service_type = resource['Compliance'].get('service_type', 'other')
                required_tags_set.update(self._get_required_tags_for_service(service_type))
        
        required_tags = sorted(required_tags_set)
        
        # Build headers
        headers = ['Service', 'ResourceType', 'ResourceId', 'ResourceName', 'State']
        headers.extend(required_tags)
        if self.tag_policy:
            headers.extend(['RequiredCount', 'PresentCount', 'MissingCount', 'ComplianceScore'])
        
        # Build data rows
        data_rows = []
        for resource in resources:
            normalized_tags = {self.report_builder.normalize_tag_key(k): v 
                             for k, v in resource.get('Tags', {}).items()}
            
            row = [
                resource.get('Service', ''),
                resource.get('ResourceType', ''),
                resource.get('ResourceId', ''),
                resource.get('ResourceName', ''),
                resource.get('State', '')
            ]
            
            # Add required tag values or MISSING
            service_type = resource.get('Compliance', {}).get('service_type', 'other') if 'Compliance' in resource else 'other'
            resource_required = self._get_required_tags_for_service(service_type)
            
            missing_count = 0
            present_count = 0
            
            for tag_key in required_tags:
                if tag_key in resource_required:
                    value = normalized_tags.get(tag_key, '')
                    if value:
                        row.append(value)
                        present_count += 1
                    else:
                        row.append('MISSING')
                        missing_count += 1
                else:
                    # Not required for this service
                    row.append('N/A')
            
            # Add summary columns
            if self.tag_policy:
                total_required = len(resource_required)
                row.append(total_required)
                row.append(present_count)
                row.append(missing_count)
                compliance = (present_count / total_required * 100) if total_required > 0 else 100.0
                row.append(f"{compliance:.1f}%")
            
            data_rows.append(row)
        
        self._write_sheet_with_auto_width(ws, headers, data_rows)
    
    def _create_recommended_tags_sheet(self, wb: Workbook, resources: List[Dict[str, Any]], sheet_name: str = "Recommended Tags") -> None:
        """Create sheet showing recommended tags with MISSING markers"""
        ws = wb.create_sheet(title=sheet_name[:31])
        
        # Collect all recommended tag columns across all resources
        recommended_tags_set = set()
        for resource in resources:
            if 'Compliance' in resource:
                service_type = resource['Compliance'].get('service_type', 'other')
                recommended_tags_set.update(self._get_recommended_tags_for_service(service_type))
        
        recommended_tags = sorted(recommended_tags_set)
        
        # Build headers
        headers = ['Service', 'ResourceType', 'ResourceId', 'ResourceName', 'State']
        headers.extend(recommended_tags)
        if self.tag_policy:
            headers.extend(['RecommendedCount', 'PresentCount', 'MissingCount', 'Coverage'])
        
        # Build data rows
        data_rows = []
        for resource in resources:
            normalized_tags = {self.report_builder.normalize_tag_key(k): v 
                             for k, v in resource.get('Tags', {}).items()}
            
            row = [
                resource.get('Service', ''),
                resource.get('ResourceType', ''),
                resource.get('ResourceId', ''),
                resource.get('ResourceName', ''),
                resource.get('State', '')
            ]
            
            # Add recommended tag values or MISSING
            service_type = resource.get('Compliance', {}).get('service_type', 'other') if 'Compliance' in resource else 'other'
            resource_recommended = self._get_recommended_tags_for_service(service_type)
            
            missing_count = 0
            present_count = 0
            
            for tag_key in recommended_tags:
                if tag_key in resource_recommended:
                    value = normalized_tags.get(tag_key, '')
                    if value:
                        row.append(value)
                        present_count += 1
                    else:
                        row.append('MISSING')
                        missing_count += 1
                else:
                    # Not recommended for this service
                    row.append('N/A')
            
            # Add summary columns
            if self.tag_policy:
                total_recommended = len(resource_recommended)
                row.append(total_recommended)
                row.append(present_count)
                row.append(missing_count)
                coverage = (present_count / total_recommended * 100) if total_recommended > 0 else 100.0
                row.append(f"{coverage:.1f}%")
            
            data_rows.append(row)
        
        self._write_sheet_with_auto_width(ws, headers, data_rows)
    
    def _create_current_tags_sheet(self, wb: Workbook, resources: List[Dict[str, Any]], sheet_name: str = "Current Tags") -> None:
        """Create sheet showing all current tags (original behavior)"""
        ws = wb.create_sheet(title=sheet_name[:31])
        
        # Collect all tag columns
        tag_cols = self._collect_tag_columns(resources)
        headers = self._build_csv_headers(tag_cols)
        
        # Build data rows
        data_rows = []
        for resource in resources:
            row = self._build_csv_row(resource, tag_cols)
            data_rows.append(row)
        
        self._write_sheet_with_auto_width(ws, headers, data_rows)
    
    # ========================================================================
    # CONSOLE OUTPUT
    # ========================================================================
    
    def print_summary(self, summary: Dict[str, Any]) -> None:
        """Print a formatted summary to console"""
        self.logger.info("\n" + "="*60)
        self.logger.info("AWS TAGGING SUMMARY REPORT")
        self.logger.info("="*60)

        self.logger.info("Overall Statistics:")
        self.logger.info(f"  Total Resources: {summary['total_resources']}")
        self.logger.info(f"  Resources with Tags: {summary['resources_with_tags']}")
        self.logger.info(f"  Resources without Tags: {summary['resources_without_tags']}")

        if summary['total_resources'] > 0:
            compliance_rate = (summary['resources_with_tags'] / summary['total_resources']) * 100
            self.logger.info(f"  Tag Compliance Rate: {compliance_rate:.1f}%")

        # Show policy compliance if available
        if self.tag_policy and summary['compliance']['compliant'] + summary['compliance']['non_compliant'] > 0:
            self.logger.info(f"Tag Policy Compliance ({self.tag_policy.get('name', 'Policy')}):")
            self.logger.info(f"  Compliant Resources: {summary['compliance']['compliant']}")
            self.logger.info(f"  Non-Compliant Resources: {summary['compliance']['non_compliant']}")
            self.logger.info(f"  Policy Compliance Rate: {summary['compliance']['compliance_rate']:.1f}%")

        self.logger.info("By Service:")
        for service, stats in summary['by_service'].items():
            compliance = (stats['with_tags'] / stats['total'] * 100) if stats['total'] > 0 else 0
            output = f"  {service}: {stats['total']} total, {stats['with_tags']} tagged ({compliance:.1f}%)"

            if self.tag_policy and 'compliant' in stats:
                policy_comp = (stats['compliant'] / stats['total'] * 100) if stats['total'] > 0 else 0
                output += f", {stats['compliant']} policy-compliant ({policy_comp:.1f}%)"

            self.logger.info(output)

        self.logger.info("Most Common Tags:")
        for tag, count in list(summary['common_tags'].items())[:10]:
            self.logger.info(f"  {tag}: {count} resources")

        self.logger.info("\n" + "="*60)
    
    def _print_resource_header(self, resource: Dict[str, Any]) -> None:
        """Print resource header information"""
        self.logger.info(f"[{resource['Service']} - {resource['ResourceType']}]")
        self.logger.info(f"   ID: {resource['ResourceId']}")
        self.logger.info(f"   Name: {resource['ResourceName']}")
        self.logger.info(f"   State: {resource['State']}")
        self.logger.info(f"   Tags ({resource['TagCount']}):")
    
    def _print_resource_tags(self, resource: Dict[str, Any]) -> None:
        """Print resource tags or no-tags message"""
        if resource['Tags']:
            for key, value in resource['Tags'].items():
                self.logger.info(f"     {key}: {value}")
        else:
            self.logger.info("     [No tags found]")
    
    def _print_compliance_status(self, compliance: Dict[str, Any]) -> None:
        """Print compliance status and details"""
        if compliance['compliant']:
            self.logger.info(f"   [COMPLIANT] Policy Compliant (Score: {compliance['compliance_score']:.1f}%)")
        else:
            self.logger.info(f"   [NON-COMPLIANT] Policy Non-Compliant (Score: {compliance['compliance_score']:.1f}%)")

            if compliance['missing_required']:
                self.logger.info(f"      Missing Required: {', '.join(compliance['missing_required'])}")

            if compliance['missing_recommended']:
                self.logger.info(f"      Missing Recommended: {', '.join(compliance['missing_recommended'])}")

            if compliance['invalid_values']:
                self.logger.info("      Invalid Tag Values:")
                for inv in compliance['invalid_values']:
                    self.logger.info(f"        - {inv['key']}: '{inv['value']}' (allowed: {', '.join(inv['allowed'])})")
    
    def print_detailed_report(self, resources: List[Dict[str, Any]]) -> None:
        """Print detailed resource information"""
        print("\n" + "="*80)
        print("DETAILED RESOURCE REPORT")
        print("="*80)
        
        for resource in resources:
            self._print_resource_header(resource)
            self._print_resource_tags(resource)
            
            # Show compliance status if policy is loaded
            if self.tag_policy and 'Compliance' in resource:
                self._print_compliance_status(resource['Compliance'])
