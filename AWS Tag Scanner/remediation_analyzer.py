"""
Tag Remediation Analyzer

Analyzes tag gaps and generates intelligent remediation recommendations.
Can work standalone with CSV/XLSX files or be integrated into the scanner.
"""

import csv
import json
import logging
from typing import Dict, List, Set, Optional, Any
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


class RemediationAnalyzer:
    """Analyzes tag gaps and generates remediation recommendations"""
    
    def __init__(self, policy: Optional[Dict] = None):
        """
        Initialize remediation analyzer
        
        Args:
            policy: Tag policy dictionary (optional)
        """
        self.policy = policy
        self.remediation_plan = []
        self.logger = logging.getLogger(__name__)
    
    def analyze_from_file(self, file_path: str) -> List[Dict]:
        """
        Analyze tags from CSV or XLSX file and generate remediation plan
        
        Args:
            file_path: Path to CSV or XLSX file from scanner output
            
        Returns:
            List of remediation recommendations
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        # Load resources based on file type
        if file_path.suffix.lower() == '.csv':
            resources = self._load_from_csv(file_path)
        elif file_path.suffix.lower() in ['.xlsx', '.xls']:
            resources = self._load_from_xlsx(file_path)
        else:
            raise ValueError(f"Unsupported file format: {file_path.suffix}")
        
        return self.analyze_resources(resources)
    
    def analyze_resources(self, resources: List[Dict]) -> List[Dict]:
        """
        Analyze resources and generate remediation plan
        
        Args:
            resources: List of resource dictionaries
            
        Returns:
            List of remediation recommendations
        """
        self.remediation_plan = []
        
        for resource in resources:
            service_type = self._get_service_type(resource)
            tags = self._extract_tags(resource)
            missing_tags = self._calculate_missing_tags(tags, service_type)
            
            if missing_tags:
                recommendations = self._generate_recommendations(resource, tags, missing_tags)
                if recommendations:
                    self.remediation_plan.append({
                        'service': service_type,
                        'resource_type': resource.get('ResourceType', resource.get('Type', 'Unknown')),
                        'resource_id': resource.get('ResourceId', resource.get('ID', 'Unknown')),
                        'resource_name': resource.get('ResourceName', tags.get('Name', 'Unnamed')),
                        'state': resource.get('State', 'Unknown'),
                        'existing_tags': len(tags),
                        'missing_count': len(missing_tags),
                        'recommendations': recommendations
                    })
        
        self.logger.info(f"Generated remediation plan for {len(self.remediation_plan)} resources")
        return self.remediation_plan
    
    def _load_from_csv(self, file_path: Path) -> List[Dict]:
        """Load resources from CSV file"""
        resources = []
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            resources = list(reader)
        self.logger.info(f"Loaded {len(resources)} resources from CSV")
        return resources
    
    def _load_from_xlsx(self, file_path: Path) -> List[Dict]:
        """Load resources from XLSX file"""
        wb = load_workbook(file_path, read_only=True)
        
        # Try to find the main data sheet
        sheet_names = wb.sheetnames
        data_sheet = None
        
        # Prioritize sheets with resource data
        priority_names = ['All Resources', 'Resources', 'EC2', 'S3', 'RDS']
        for name in priority_names:
            if name in sheet_names:
                data_sheet = wb[name]
                break
        
        if not data_sheet:
            data_sheet = wb.active
        
        # Convert sheet to list of dictionaries
        resources = []
        headers = []
        for i, row in enumerate(data_sheet.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(cell) if cell else f"Column{i}" for i, cell in enumerate(row)]
            else:
                resource = {}
                for header, value in zip(headers, row):
                    resource[header] = str(value) if value is not None else ''
                resources.append(resource)
        
        self.logger.info(f"Loaded {len(resources)} resources from XLSX sheet '{data_sheet.title}'")
        wb.close()
        return resources
    
    def _get_service_type(self, resource: Dict) -> str:
        """Determine service type from resource"""
        service = resource.get('Service', '').lower()
        resource_type = resource.get('ResourceType', '').lower()
        
        if service == 'ec2':
            if 'instance' in resource_type:
                return 'ec2'
            elif 'volume' in resource_type:
                return 'ebs'
        elif service == 's3':
            return 's3'
        elif service == 'rds':
            return 'rds'
        elif service == 'lambda':
            return 'lambda'
        elif service == 'elb':
            return 'elb'
        
        return 'other'
    
    def _extract_tags(self, resource: Dict) -> Dict:
        """Extract tags from resource dictionary"""
        tags = {}
        for key, value in resource.items():
            # Scanner exports tags with 'Tag_' prefix or as column names
            if key.startswith('Tag_'):
                tag_name = key[4:]  # Remove 'Tag_' prefix
                tags[tag_name] = value
            elif key.startswith('tag_'):
                tag_name = key[4:]
                tags[tag_name] = value
            # Also check if it's a known tag name directly
            elif key in ['Name', 'customer', 'department', 'environment', 'application',
                        'servicescope', 'rebill', 'function', 'data-priority',
                        'technical-contact', 'wafv2', 'backup', 'critical-list',
                        'criticality', 'Patch Group']:
                tags[key] = value
        return tags
    
    def _get_required_tags_for_service(self, service_type: str) -> List[str]:
        """Get required tags for a specific service from policy"""
        if not self.policy:
            return []
        
        # Core tags apply to all services
        core_tags = [tag['key'] for tag in self.policy.get('core_tags', {}).get('tags', [])]
        
        # Service-specific required tags
        service_reqs = self.policy.get('service_specific_requirements', {}).get(service_type, {})
        service_tags = service_reqs.get('required_tags', [])
        
        return core_tags + service_tags
    
    def _calculate_missing_tags(self, tags: Dict, service_type: str = None) -> Set[str]:
        """Calculate which required tags are missing"""
        if self.policy and service_type:
            required = set(self._get_required_tags_for_service(service_type))
        else:
            # Fallback to common tags if no policy
            required = {
                'Name', 'customer', 'department', 'environment', 'application',
                'servicescope', 'rebill', 'function', 'data-priority',
                'technical-contact', 'wafv2', 'backup', 'critical-list',
                'criticality', 'Patch Group'
            }
        
        existing = set(k for k, v in tags.items() if v)
        return required - existing
    
    def _generate_recommendations(self, resource: Dict, tags: Dict, missing: Set[str]) -> List[Dict]:
        """Generate smart default values based on existing tags"""
        recommendations = []
        
        # Get existing tag values for reference
        environment = tags.get('environment', '').lower()
        function = tags.get('function', '').lower()
        department = tags.get('department', '')
        critical_list = tags.get('critical-list', 'N')
        
        # Generate recommendations for each missing tag
        for missing_tag in missing:
            rec = {'tag': missing_tag, 'value': None, 'logic': ''}
            
            if missing_tag == 'customer':
                rec['value'] = department if department else 'UNKNOWN'
                rec['logic'] = 'Copied from department tag'
            
            elif missing_tag == 'servicescope':
                if environment == 'prod':
                    rec['value'] = 'University-Wide'
                    rec['logic'] = 'Production resource → University-Wide'
                else:
                    rec['value'] = 'Dept-Internal'
                    rec['logic'] = 'Non-production → Dept-Internal'
            
            elif missing_tag == 'rebill':
                rec['value'] = 'N'
                rec['logic'] = 'Default: No chargeback (change if MOA exists)'
            
            elif missing_tag == 'data-priority':
                if function in ['db', 'database']:
                    rec['value'] = 'pii'
                    rec['logic'] = 'Database → PII (verify classification)'
                elif environment == 'prod':
                    rec['value'] = 'non-sensitive'
                    rec['logic'] = 'Production → non-sensitive (verify)'
                else:
                    rec['value'] = 'non-sensitive'
                    rec['logic'] = 'Default: non-sensitive'
            
            elif missing_tag == 'wafv2':
                if function == 'web':
                    rec['value'] = 'default'
                    rec['logic'] = 'Web server → default WAF rules'
                elif function in ['app', 'db', 'database']:
                    rec['value'] = 'exclude-all'
                    rec['logic'] = 'App/DB server → exclude from WAF'
                else:
                    rec['value'] = 'exclude-all'
                    rec['logic'] = 'Default: exclude-all'
            
            elif missing_tag == 'backup':
                if environment == 'prod':
                    rec['value'] = 'Yes'
                    rec['logic'] = 'Production → backup enabled'
                elif critical_list == 'Y':
                    rec['value'] = 'Yes'
                    rec['logic'] = 'Critical resource → backup enabled'
                else:
                    rec['value'] = 'No'
                    rec['logic'] = 'Non-production/non-critical → no backup'
            
            elif missing_tag == 'criticality':
                if critical_list == 'Y':
                    rec['value'] = '1 - Mission Critical'
                    rec['logic'] = 'On critical list → Mission Critical'
                elif environment == 'prod':
                    rec['value'] = '2 - Business Critical'
                    rec['logic'] = 'Production → Business Critical'
                else:
                    rec['value'] = '3 - Operational Support'
                    rec['logic'] = 'Default: Operational Support'
            
            elif missing_tag == 'Patch Group':
                if environment in ['prod', 'test', 'dev']:
                    rec['value'] = environment
                    rec['logic'] = 'Copied from environment tag'
                else:
                    rec['value'] = 'dev'
                    rec['logic'] = 'Default: dev patch group'
            
            elif missing_tag == 'technical-contact':
                rec['value'] = 'REQUIRED-CONTACT@example.com'
                rec['logic'] = 'PLACEHOLDER - Must be updated with actual contact'
            
            elif missing_tag == 'Name':
                rec['value'] = resource.get('ResourceId', 'unnamed')
                rec['logic'] = 'Using resource ID as name'
            
            elif missing_tag == 'department':
                rec['value'] = 'UNKNOWN'
                rec['logic'] = 'PLACEHOLDER - Must be set manually'
            
            elif missing_tag == 'environment':
                rec['value'] = 'dev'
                rec['logic'] = 'DEFAULT - Verify actual environment'
            
            elif missing_tag == 'application':
                rec['value'] = 'UNKNOWN'
                rec['logic'] = 'PLACEHOLDER - Must be set manually'
            
            elif missing_tag == 'function':
                rec['value'] = 'app'
                rec['logic'] = 'DEFAULT - Verify actual function'
            
            elif missing_tag == 'critical-list':
                rec['value'] = 'N'
                rec['logic'] = 'Default: Not on critical list'
            
            if rec['value']:
                recommendations.append(rec)
        
        return recommendations
    
    def get_gap_analysis(self) -> Dict[str, Dict[str, int]]:
        """
        Get summary of tag gaps by service type
        
        Returns:
            Dictionary mapping service types to tag gaps
        """
        service_stats = defaultdict(lambda: defaultdict(int))
        
        for item in self.remediation_plan:
            service = item['service']
            for rec in item['recommendations']:
                service_stats[service][rec['tag']] += 1
        
        return dict(service_stats)
    
    def export_to_csv(self, output_file: str):
        """Export remediation plan to CSV"""
        from security_utils import sanitize_csv_value
        
        rows = []
        for item in self.remediation_plan:
            for rec in item['recommendations']:
                rows.append({
                    'Service': sanitize_csv_value(item['service'].upper()),
                    'ResourceType': sanitize_csv_value(item['resource_type']),
                    'ResourceId': sanitize_csv_value(item['resource_id']),
                    'ResourceName': sanitize_csv_value(item['resource_name']),
                    'State': sanitize_csv_value(item['state']),
                    'TagName': sanitize_csv_value(rec['tag']),
                    'RecommendedValue': sanitize_csv_value(rec['value']),
                    'Logic': sanitize_csv_value(rec['logic']),
                    'ExistingTagCount': item['existing_tags'],
                    'MissingTagCount': item['missing_count']
                })
        
        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            if rows:
                writer = csv.DictWriter(f, fieldnames=rows[0].keys())
                writer.writeheader()
                writer.writerows(rows)
        
        self.logger.info(f"Exported {len(rows)} tag recommendations to {output_file}")
        return len(rows)
    
    def export_to_json(self, output_file: str):
        """Export remediation plan to JSON"""
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(self.remediation_plan, f, indent=2)
        
        self.logger.info(f"Exported {len(self.remediation_plan)} remediation plans to {output_file}")
    
    def print_summary(self):
        """Print remediation summary to console"""
        print("\n" + "=" * 80)
        print("REMEDIATION SUMMARY")
        if self.policy:
            print("   (Using service-specific requirements)")
        print("=" * 80)
        
        # Group by service type
        by_service = defaultdict(list)
        for item in self.remediation_plan:
            service = item.get('service', 'unknown')
            by_service[service].append(item)
        
        for service_type, items in sorted(by_service.items()):
            print(f"\n[{service_type.upper()}] Service: {len(items)} resources need remediation")
            
            # Show a few examples
            for item in items[:3]:
                print(f"\n  Resource: {item['resource_name']}")
                print(f"  ID: {item['resource_id']}")
                print(f"  Type: {item['resource_type']}")
                print(f"  State: {item.get('state', 'Unknown')}")
                print(f"  Existing tags: {item['existing_tags']}, Missing: {item['missing_count']}")
                print(f"  Recommendations:")
                for rec in item['recommendations'][:5]:  # Show first 5
                    print(f"    • {rec['tag']:20s} = {rec['value']:30s}  # {rec['logic']}")
                if len(item['recommendations']) > 5:
                    more_count = len(item['recommendations']) - 5
                    print(f"    ... and {more_count} more tags")
            
            if len(items) > 3:
                remaining = len(items) - 3
                print(f"\n  ... and {remaining} more {service_type} resources")
