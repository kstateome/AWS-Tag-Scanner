#!/usr/bin/env python3
"""
AWS Tag Scanner
Scans AWS resources across multiple services to identify their tags and generate reports
for compliance verification.

Requires `boto3` (and `botocore`) and AWS CLI to be configured and authenticated.
This script performs read-only operations against AWS and does not modify resources.

Perfect for use in AWS CloudShell or local environments with `boto3` installed.
"""

import boto3
import json
import csv
import os
import logging
from datetime import datetime
from botocore.exceptions import ClientError, NoCredentialsError
import argparse
import sys
from typing import Dict, List, Any
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ============================================================================
# SECURITY UTILITIES
# ============================================================================

def validate_file_path(file_path: str, max_size_mb: int = 100) -> str:
    """
    Validate file path to prevent path traversal attacks.

    @param file_path: Path to validate
    @param max_size_mb: Maximum allowed file size in MB
    @return: Absolute validated path
    @raises ValueError: If path is invalid, missing, not a file, or too large
    """
    if not file_path:
        raise ValueError("File path cannot be empty")

    # Convert to absolute path
    abs_path = os.path.abspath(file_path)

    # Check if file exists
    if not os.path.exists(abs_path):
        raise ValueError(f"File not found: {file_path}")

    # Check if it's a file (not directory)
    if not os.path.isfile(abs_path):
        raise ValueError(f"Path is not a file: {file_path}")

    # Check file size
    file_size = os.path.getsize(abs_path)
    max_bytes = max_size_mb * 1024 * 1024
    if file_size > max_bytes:
        raise ValueError(f"File too large: {file_size} bytes (max: {max_bytes})")

    return abs_path


def validate_output_filename(filename: str) -> str:
    """
    Validate a user-supplied output filename (no path traversal, no absolute paths).

    - Disallow path separators and absolute paths
    - Disallow parent-directory segments
    Returns the sanitized filename (as a string)
    Raises ValueError on invalid names
    """
    if not filename:
        raise ValueError("Filename cannot be empty")

    p = Path(filename)

    # Disallow absolute paths
    if p.is_absolute():
        raise ValueError("Absolute paths are not allowed for output filename")

    # Disallow parent directory traversal
    if '..' in p.parts:
        raise ValueError("Parent directory traversal is not allowed in filename")

    # Disallow any path separators by ensuring it has only a name
    if len(p.parts) != 1:
        raise ValueError("Filename must not contain path separators")

    # Basic cleanliness: remove surrounding whitespace
    clean = p.name.strip()
    if not clean:
        raise ValueError("Filename must contain visible characters")

    return clean


def load_json_safely(file_path: str, max_size_mb: int = 10) -> Dict[str, Any]:
    """
    Load JSON file with security validations
    
    Args:
        file_path: Path to JSON file
        max_size_mb: Maximum allowed file size in MB
        
    Returns:
        Parsed JSON data
        
    Raises:
        ValueError: If file is invalid or too large
    """
    # Validate file path
    validated_path = validate_file_path(file_path, max_size_mb)
    
    # Load and parse JSON
    try:
        with open(validated_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except json.JSONDecodeError as e:
        raise ValueError(f"Invalid JSON format: {e}")
    except UnicodeDecodeError as e:
        raise ValueError(f"Invalid file encoding: {e}")
    
    if not isinstance(data, dict):
        raise ValueError("JSON root must be an object")
    
    return data


def sanitize_csv_value(value: Any) -> Any:
    """
    Sanitize value to prevent CSV injection attacks
    
    Prefixes dangerous characters with single quote to prevent
    formula execution in Excel/LibreOffice/Google Sheets
    
    Args:
        value: Value to sanitize
        
    Returns:
        Sanitized value safe for CSV export
    """
    if not value or not isinstance(value, str):
        return value
    
    # Characters that could trigger formula execution
    dangerous_chars = ['=', '+', '-', '@', '\t', '\r', '\n']
    
    # If value starts with dangerous character, prefix with single quote
    if value and value[0] in dangerous_chars:
        return "'" + value
    
    return value


class AWSTagScanner:
    def __init__(self, profile: str | None = None, region: str | None = None, dry_run: bool = False, policy_file: str | None = None):
        """Initialize AWS clients for various services"""
        self.dry_run = dry_run
        self.tag_policy: Dict[str, Any] | None = None
        self.region: str = ""
        # module logger
        self.logger = logging.getLogger(__name__)
        
        # Load tag policy if provided
        if policy_file:
            self.load_tag_policy(policy_file)
        
        if self.dry_run:
            self.logger.info("[INFO] DRY RUN MODE: Will only check authentication, not scan resources")
        
        try:
            self.session = boto3.Session(profile_name=profile, region_name=region)
            self.region = region or self.session.region_name or "us-east-1"

            # Test authentication first
            sts = self.session.client('sts')
            identity = sts.get_caller_identity()
            # Avoid printing full ARN in logs; show account id only by default
            account = identity.get('Account', 'Unknown')
            self.logger.info(f"[OK] Authenticated into account: {account}")
            self.logger.info(f"[OK] Connected to AWS in region: {self.region}")
            
            if self.dry_run:
                self.logger.info("[OK] Authentication successful - dry run complete")
                return
                
            # Initialize clients for different AWS services
            self.ec2 = self.session.client('ec2')
            self.s3 = self.session.client('s3')
            self.rds = self.session.client('rds')
            self.lambda_client = self.session.client('lambda')
            self.elbv2 = self.session.client('elbv2')
            self.elb = self.session.client('elb')
            
        except NoCredentialsError:
            self.logger.error("[ERROR] AWS credentials not found. Please configure AWS CLI first.")
            self.logger.error("[TIP] For SAML accounts, use AWS CloudShell for automatic authentication.")
            sys.exit(1)
        except Exception as e:
            self.logger.error(f"[ERROR] Error initializing AWS clients: {e}")
            sys.exit(1)
    
    def load_tag_policy(self, policy_file: str) -> None:
        """Load service-specific tag policy from JSON file with security validations"""
        try:
            # Secure JSON loading with size validation
            policy_data = load_json_safely(policy_file, max_size_mb=10)
            self.tag_policy = policy_data.get('tag_policy', {})
            
            # Validate it's a service-specific policy
            if not self.tag_policy or 'service_specific_requirements' not in self.tag_policy:
                self.logger.warning("Policy file doesn't contain service-specific requirements")
                self.logger.warning("Expected 'service_specific_requirements' section")
                self.tag_policy = None
                return
            self.logger.info(f"Loaded service-specific tag policy: {self.tag_policy.get('name', 'Unknown')}")
            services = list(self.tag_policy.get('service_specific_requirements', {}).keys())
            self.logger.info(f"Services configured: {', '.join(services)}")

            # Show core tags
            core_tags = self.tag_policy.get('core_tags', {}).get('tags', [])
            self.logger.info(f"Core tags (all services): {len(core_tags)}")
                
        except (FileNotFoundError, json.JSONDecodeError, ValueError) as e:
            # load_json_safely may raise ValueError for invalid files/paths/sizes
            self.logger.warning(f"Error loading tag policy '{policy_file}': {e}")
            self.tag_policy = None
    
    def normalize_tag_key(self, tag_key: str) -> str:
        """Normalize tag key using aliases from policy"""
        if not self.tag_policy:
            return tag_key
        
        aliases = self.tag_policy.get('tag_aliases', {})
        return aliases.get(tag_key, tag_key)
    
    def get_service_type_from_resource(self, resource: Dict[str, Any]) -> str:
        """Map resource service/type to policy service type (returns uppercase to match policy JSON)"""
        service = resource.get('Service', '')
        resource_type = resource.get('ResourceType', '')
        
        # Map AWS service names to policy service types (uppercase to match policy JSON keys)
        if service == 'EC2':
            if 'Instance' in resource_type:
                return 'EC2'
            elif 'Volume' in resource_type:
                return 'EBS'
        elif service == 'S3':
            return 'S3'
        elif service == 'RDS':
            return 'RDS'
        elif service == 'Lambda':
            return 'Lambda'
        elif service == 'ELB':
            return 'ELB'
        
        return 'other'
    
    def check_resource_compliance(self, resource: Dict[str, Any]) -> Dict[str, Any]:
        """Check resource compliance using service-specific requirements"""
        if not self.tag_policy:
            return {
                'compliant': None,
                'missing_required': [],
                'missing_recommended': [],
                'unneeded_tags': [],
                'invalid_values': [],
                'compliance_score': 100.0
            }
        
        # Get service type for this resource
        service_type = self.get_service_type_from_resource(resource)
        
        # Get service-specific requirements
        service_reqs = self.tag_policy.get('service_specific_requirements', {}).get(service_type)
        core_tags = self.tag_policy.get('core_tags', {}).get('tags', [])
        
        # Build required and recommended tag lists
        required_tag_keys = [tag['key'] for tag in core_tags]
        recommended_tag_keys = []
        
        if service_reqs:
            required_tag_keys.extend(service_reqs.get('required_tags', []))
            recommended_tag_keys.extend(service_reqs.get('recommended_tags', []))
        
        # Build set of all policy-defined tags (required + recommended)
        all_policy_tags = set(required_tag_keys + recommended_tag_keys)
        
        # Get resource tags and normalize
        tags = resource.get('Tags', {})
        normalized_tags = {self.normalize_tag_key(k): v for k, v in tags.items()}
        
        missing_required = []
        missing_recommended = []
        unneeded_tags = []
        
        # Check required tags
        for tag_key in required_tag_keys:
            if tag_key not in normalized_tags or not normalized_tags[tag_key]:
                missing_required.append(tag_key)
        
        # Check recommended tags
        for tag_key in recommended_tag_keys:
            if tag_key not in normalized_tags or not normalized_tags[tag_key]:
                missing_recommended.append(tag_key)
        
        # Check for unneeded tags (tags not in policy)
        for tag_key in normalized_tags.keys():
            if tag_key not in all_policy_tags:
                unneeded_tags.append(tag_key)
        
        # Calculate compliance
        total_required = len(required_tag_keys)
        total_recommended = len(recommended_tag_keys)
        total_tags = total_required + total_recommended
        
        if total_tags > 0:
            found_required = total_required - len(missing_required)
            found_recommended = total_recommended - len(missing_recommended)
            compliance_score = ((found_required + found_recommended) / total_tags) * 100
        else:
            compliance_score = 100.0
        
        return {
            'compliant': len(missing_required) == 0,
            'missing_required': missing_required,
            'missing_recommended': missing_recommended,
            'unneeded_tags': sorted(unneeded_tags),
            'invalid_values': [],
            'compliance_score': compliance_score,
            'service_type': service_type,
            'total_required': total_required,
            'total_recommended': total_recommended
        }
    
    def scan_ec2_instances(self) -> List[Dict[str, Any]]:
        """Scan EC2 instances and their tags"""
        resources = []
        try:
            self.logger.debug("    [API] Calling describe_instances (read-only)...")
            paginator = self.ec2.get_paginator('describe_instances')
            for page in paginator.paginate():
                for reservation in page.get('Reservations', []):
                    for instance in reservation.get('Instances', []):
                        tags = {tag['Key']: tag['Value'] for tag in instance.get('Tags', [])}
                        resources.append({
                            'Service': 'EC2',
                            'ResourceType': 'Instance',
                            'ResourceId': instance['InstanceId'],
                            'ResourceName': tags.get('Name', ''),
                            'State': instance.get('State', {}).get('Name', ''),
                            'Tags': tags,
                            'TagCount': len(tags)
                        })
        except ClientError as e:
            if e.response['Error']['Code'] == 'UnauthorizedOperation':
                self.logger.warning("[WARN] No permission to scan EC2 instances")
            else:
                self.logger.warning(f"[WARN] Error scanning EC2 instances: {e}")
        
        return resources
    
    def scan_ebs_volumes(self) -> List[Dict[str, Any]]:
        """Scan EBS volumes and their tags"""
        resources = []
        try:
            self.logger.debug("    [API] Calling describe_volumes (read-only)...")
            paginator = self.ec2.get_paginator('describe_volumes')
            for page in paginator.paginate():
                for volume in page.get('Volumes', []):
                    tags = {tag['Key']: tag['Value'] for tag in volume.get('Tags', [])}
                    resources.append({
                        'Service': 'EC2',
                        'ResourceType': 'EBS Volume',
                        'ResourceId': volume['VolumeId'],
                        'ResourceName': tags.get('Name', ''),
                        'State': volume.get('State', ''),
                        'Tags': tags,
                        'TagCount': len(tags)
                    })
        except ClientError as e:
            if e.response['Error']['Code'] == 'UnauthorizedOperation':
                self.logger.warning("[WARN] No permission to scan EBS volumes")
            else:
                self.logger.warning(f"[WARN] Error scanning EBS volumes: {e}")
        
        return resources
    
    def scan_s3_buckets(self) -> List[Dict[str, Any]]:
        """Scan S3 buckets and their tags"""
        resources = []
        try:
            self.logger.debug("    [API] Calling list_buckets (read-only)...")
            response = self.s3.list_buckets()
            for bucket in response['Buckets']:
                bucket_name = bucket['Name']
                tags = {}
                try:
                    tag_response = self.s3.get_bucket_tagging(Bucket=bucket_name)
                    tags = {tag['Key']: tag['Value'] for tag in tag_response.get('TagSet', [])}
                except ClientError:
                    # Bucket has no tags or access denied
                    pass
                
                resources.append({
                    'Service': 'S3',
                    'ResourceType': 'Bucket',
                    'ResourceId': bucket_name,
                    'ResourceName': bucket_name,
                    'State': 'Active',
                    'Tags': tags,
                    'TagCount': len(tags)
                })
        except ClientError as e:
            if e.response['Error']['Code'] == 'AccessDenied':
                self.logger.warning("No permission to scan S3 buckets")
            else:
                self.logger.warning(f"Error scanning S3 buckets: {e}")
        
        return resources
    
    def scan_rds_instances(self) -> List[Dict[str, Any]]:
        """Scan RDS instances and their tags"""
        resources = []
        try:
            self.logger.debug("    [API] Calling describe_db_instances (read-only)...")
            paginator = self.rds.get_paginator('describe_db_instances')
            for page in paginator.paginate():
                for db in page.get('DBInstances', []):
                    tags = {}
                    try:
                        tag_response = self.rds.list_tags_for_resource(
                            ResourceName=db['DBInstanceArn']
                        )
                        tags = {tag['Key']: tag['Value'] for tag in tag_response.get('TagList', [])}
                    except ClientError:
                        pass

                    resources.append({
                        'Service': 'RDS',
                        'ResourceType': 'DB Instance',
                        'ResourceId': db['DBInstanceIdentifier'],
                        'ResourceName': db['DBInstanceIdentifier'],
                        'State': db['DBInstanceStatus'],
                        'Tags': tags,
                        'TagCount': len(tags)
                    })
        except ClientError as e:
            if e.response['Error']['Code'] == 'AccessDenied':
                self.logger.warning("No permission to scan RDS instances")
            else:
                self.logger.warning(f"Error scanning RDS instances: {e}")
        
        return resources
    
    def scan_lambda_functions(self) -> List[Dict[str, Any]]:
        """Scan Lambda functions and their tags"""
        resources = []
        try:
            self.logger.debug("    [API] Calling list_functions (read-only)...")
            paginator = self.lambda_client.get_paginator('list_functions')
            for page in paginator.paginate():
                for function in page.get('Functions', []):
                    tags = {}
                    try:
                        tag_response = self.lambda_client.list_tags(
                            Resource=function['FunctionArn']
                        )
                        tags = tag_response.get('Tags', {})
                    except ClientError:
                        pass

                    resources.append({
                        'Service': 'Lambda',
                        'ResourceType': 'Function',
                        'ResourceId': function['FunctionName'],
                        'ResourceName': function['FunctionName'],
                        'State': function.get('State', 'Active'),
                        'Tags': tags,
                        'TagCount': len(tags)
                    })
        except ClientError as e:
            if e.response['Error']['Code'] == 'AccessDenied':
                self.logger.warning("No permission to scan Lambda functions")
            else:
                self.logger.warning(f"Error scanning Lambda functions: {e}")
        
        return resources
    
    def _scan_alb_nlb(self) -> List[Dict[str, Any]]:
        """Scan Application and Network Load Balancers (ALB/NLB)"""
        resources = []
        try:
            self.logger.debug("    [API] Calling describe_load_balancers (ALB/NLB) (read-only)...")
            paginator = self.elbv2.get_paginator('describe_load_balancers')
            for page in paginator.paginate():
                for lb in page.get('LoadBalancers', []):
                    tags = {}
                    try:
                        tag_response = self.elbv2.describe_tags(
                            ResourceArns=[lb['LoadBalancerArn']]
                        )
                        if tag_response['TagDescriptions']:
                            tags = {tag['Key']: tag['Value'] 
                                   for tag in tag_response['TagDescriptions'][0].get('Tags', [])}
                    except ClientError:
                        pass

                    resources.append({
                        'Service': 'ELB',
                        'ResourceType': f"{lb['Type'].upper()} Load Balancer",
                        'ResourceId': lb['LoadBalancerName'],
                        'ResourceName': lb['LoadBalancerName'],
                        'State': lb['State']['Code'],
                        'Tags': tags,
                        'TagCount': len(tags)
                    })
        except ClientError as e:
            if e.response['Error']['Code'] == 'AccessDenied':
                self.logger.warning("No permission to scan ALB/NLB")
            else:
                self.logger.warning(f"Error scanning ALB/NLB: {e}")
        
        return resources
    
    def _scan_classic_elb(self) -> List[Dict[str, Any]]:
        """Scan Classic Load Balancers"""
        resources = []
        try:
            self.logger.debug("    [API] Calling describe_load_balancers (Classic) (read-only)...")
            paginator = self.elb.get_paginator('describe_load_balancers')
            for page in paginator.paginate():
                for lb in page.get('LoadBalancerDescriptions', []):
                    tags = {}
                    try:
                        tag_response = self.elb.describe_tags(
                            LoadBalancerNames=[lb['LoadBalancerName']]
                        )
                        if tag_response['TagDescriptions']:
                            tags = {tag['Key']: tag['Value'] 
                                   for tag in tag_response['TagDescriptions'][0].get('Tags', [])}
                    except ClientError:
                        pass

                    resources.append({
                        'Service': 'ELB',
                        'ResourceType': 'Classic Load Balancer',
                        'ResourceId': lb['LoadBalancerName'],
                        'ResourceName': lb['LoadBalancerName'],
                        'State': 'Active',
                        'Tags': tags,
                        'TagCount': len(tags)
                    })
        except ClientError as e:
            if e.response['Error']['Code'] == 'AccessDenied':
                self.logger.warning("No permission to scan Classic ELB")
            else:
                self.logger.warning(f"Error scanning Classic ELB: {e}")
        
        return resources
    
    def scan_load_balancers(self) -> List[Dict[str, Any]]:
        """Scan all types of Load Balancers (ALB/NLB and Classic)"""
        resources = []
        resources.extend(self._scan_alb_nlb())
        resources.extend(self._scan_classic_elb())
        return resources
    
    def scan_all_resources(self) -> List[Dict[str, Any]]:
        """Scan all supported AWS resources"""
        if self.dry_run:
            self.logger.info("[INFO] Dry run mode - skipping resource scan")
            return []
        self.logger.info("[SCAN] Scanning AWS resources for tags...")
        self.logger.info("[INFO] This script only performs READ operations - no changes will be made")
        
        all_resources = []
        
        print("  [EC2] Scanning EC2 instances...")
        all_resources.extend(self.scan_ec2_instances())
        
        print("  [EBS] Scanning EBS volumes...")
        all_resources.extend(self.scan_ebs_volumes())
        
        print("  [S3] Scanning S3 buckets...")
        all_resources.extend(self.scan_s3_buckets())
        
        print("  [RDS] Scanning RDS instances...")
        all_resources.extend(self.scan_rds_instances())
        
        print("  [Lambda] Scanning Lambda functions...")
        all_resources.extend(self.scan_lambda_functions())
        
        print("  [ELB] Scanning Load Balancers...")
        all_resources.extend(self.scan_load_balancers())
        
        self.logger.info(f"[OK] Scan complete! Found {len(all_resources)} resources")
        return all_resources
    
    def _calculate_service_breakdown(self, resources: List[Dict[str, Any]], summary: Dict[str, Any]) -> tuple[int, int]:
        """Calculate per-service statistics and compliance counts"""
        compliant_count = 0
        non_compliant_count = 0
        
        for resource in resources:
            service = resource['Service']
            if service not in summary['by_service']:
                summary['by_service'][service] = {
                    'total': 0,
                    'with_tags': 0,
                    'without_tags': 0,
                    'compliant': 0,
                    'non_compliant': 0
                }
            
            summary['by_service'][service]['total'] += 1
            if resource['TagCount'] > 0:
                summary['by_service'][service]['with_tags'] += 1
            else:
                summary['by_service'][service]['without_tags'] += 1
            
            # Check compliance if policy is loaded
            if self.tag_policy:
                compliance = self.check_resource_compliance(resource)
                resource['Compliance'] = compliance
                
                if compliance['compliant']:
                    compliant_count += 1
                    summary['by_service'][service]['compliant'] += 1
                else:
                    non_compliant_count += 1
                    summary['by_service'][service]['non_compliant'] += 1
        
        return compliant_count, non_compliant_count
    
    def _calculate_compliance_summary(self, summary: Dict[str, Any], compliant_count: int, 
                                     non_compliant_count: int, total_resources: int) -> None:
        """Calculate overall compliance metrics"""
        if self.tag_policy and total_resources > 0:
            summary['compliance']['compliant'] = compliant_count
            summary['compliance']['non_compliant'] = non_compliant_count
            summary['compliance']['compliance_rate'] = (compliant_count / total_resources) * 100
    
    def _analyze_common_tags(self, resources: List[Dict[str, Any]]) -> Dict[str, int]:
        """Analyze and count common tags across all resources"""
        all_tags: Dict[str, int] = {}
        for resource in resources:
            for tag_key in resource['Tags'].keys():
                all_tags[tag_key] = all_tags.get(tag_key, 0) + 1
        
        return dict(sorted(all_tags.items(), key=lambda x: x[1], reverse=True))
    
    def generate_summary_report(self, resources: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Generate a summary report of tag compliance"""
        summary: Dict[str, Any] = {
            'total_resources': len(resources),
            'resources_with_tags': len([r for r in resources if r['TagCount'] > 0]),
            'resources_without_tags': len([r for r in resources if r['TagCount'] == 0]),
            'by_service': {},
            'common_tags': {},
            'compliance': {
                'compliant': 0,
                'non_compliant': 0,
                'compliance_rate': 0.0
            }
        }
        
        # Calculate service breakdown and compliance
        compliant_count, non_compliant_count = self._calculate_service_breakdown(resources, summary)
        
        # Calculate overall compliance
        self._calculate_compliance_summary(summary, compliant_count, non_compliant_count, len(resources))
        
        # Analyze common tags
        summary['common_tags'] = self._analyze_common_tags(resources)
        
        return summary
    
    def _generate_csv_filename(self, filename: str | None = None) -> str:
        """Generate CSV filename with timestamp if not provided"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"aws_tags_scan_{timestamp}.csv"
        return filename
    
    def _collect_tag_columns(self, resources: List[Dict[str, Any]]) -> List[str]:
        """Collect all unique tag keys from resources and return sorted list"""
        all_tag_keys: set[str] = set()
        for resource in resources:
            all_tag_keys.update(resource['Tags'].keys())
        return sorted(all_tag_keys)
    
    def _build_csv_headers(self, tag_columns: List[str]) -> List[str]:
        """Build CSV header row"""
        headers = ['Service', 'ResourceType', 'ResourceId', 'ResourceName', 'State', 'TagCount']
        headers.extend(tag_columns)
        
        if self.tag_policy:
            headers.extend(['PolicyCompliant', 'ComplianceScore', 'MissingRequired', 
                          'MissingRecommended', 'UnneededTags'])
        return headers
    
    def _get_service_required_tags(self, resource: Dict[str, Any]) -> set[str]:
        """Get the set of tags required for a specific resource's service type"""
        if not self.tag_policy or 'Compliance' not in resource:
            return set()
        
        comp = resource['Compliance']
        service_type = comp.get('service_type', 'other')
        service_reqs = self.tag_policy.get('service_specific_requirements', {}).get(service_type, {})
        core_tags = self.tag_policy.get('core_tags', {}).get('tags', [])
        
        # Build set of tags that are valid for this service
        service_required_tags = {tag['key'] for tag in core_tags}
        if service_reqs:
            service_required_tags.update(service_reqs.get('required_tags', []))
            service_required_tags.update(service_reqs.get('recommended_tags', []))
        
        return service_required_tags
    
    def _format_csv_tag_value(self, tag_key: str, tag_value: str, 
                             service_required_tags: set[str]) -> str:
        """Format tag value for CSV based on policy requirements"""
        if not self.tag_policy or not service_required_tags:
            return tag_value
        
        normalized_key = self.normalize_tag_key(tag_key)
        
        if normalized_key not in service_required_tags:
            # This tag is not needed for this service type
            if tag_value:
                return f"X - {tag_value}"
            else:
                return "X"
        
        # Tag is required/recommended - return as-is (blank if missing)
        return tag_value
    
    def _build_csv_row(self, resource: Dict[str, Any], tag_columns: List[str]) -> List[Any]:
        """Build a single CSV row for a resource"""
        row: List[Any] = [
            sanitize_csv_value(resource['Service']),
            sanitize_csv_value(resource['ResourceType']), 
            sanitize_csv_value(resource['ResourceId']),
            sanitize_csv_value(resource['ResourceName']),
            sanitize_csv_value(resource['State']),
            resource['TagCount']
        ]
        
        # Get service-specific required tags
        service_required_tags = self._get_service_required_tags(resource)
        
        # Add tag values
        for tag_key in tag_columns:
            tag_value = resource['Tags'].get(tag_key, '')
            formatted_value = self._format_csv_tag_value(tag_key, tag_value, service_required_tags)
            row.append(sanitize_csv_value(formatted_value))
        
        # Add compliance columns if available
        if self.tag_policy and 'Compliance' in resource:
            comp = resource['Compliance']
            row.append('YES' if comp['compliant'] else 'NO')
            row.append(f"{comp['compliance_score']:.1f}%")
            row.append('; '.join(comp['missing_required']) if comp['missing_required'] else '')
            row.append('; '.join(comp['missing_recommended']) if comp['missing_recommended'] else '')
            row.append('; '.join(comp['unneeded_tags']) if comp['unneeded_tags'] else '')
        
        return row
    
    def _print_csv_legend(self) -> None:
        """Print CSV legend explaining tag marking system"""
        self.logger.info("* Policy compliance columns included")
        self.logger.info("* Tag column legend:")
        self.logger.info("  - Blank cell = Tag IS required but missing from resource")
        self.logger.info("  - 'X' = Tag is NOT required for this service type")
        self.logger.info("  - Value shown = Tag exists and is appropriate")
        self.logger.info("* See 'MissingRequired' column for tags that need to be added")
    
    def export_to_csv(self, resources: List[Dict[str, Any]], filename: str | None = None) -> str:
        """Export resources to CSV format"""
        # If a filename was supplied, validate it (no paths)
        if filename:
            try:
                clean = validate_output_filename(filename)
            except ValueError as e:
                raise
            # Ensure extension
            if not clean.lower().endswith('.csv'):
                clean = clean + '.csv'
            filename = clean
        else:
            filename = self._generate_csv_filename(filename)
        tag_columns = self._collect_tag_columns(resources)
        headers = self._build_csv_headers(tag_columns)
        
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(headers)
            
            # Write data rows
            for resource in resources:
                row = self._build_csv_row(resource, tag_columns)
                writer.writerow(row)
        
        self.logger.info(f"[CSV] Report exported to: {filename}")
        if self.tag_policy:
            self._print_csv_legend()
        
        return filename
    
    def export_to_json(self, resources: List[Dict[str, Any]], filename: str | None = None) -> str:
        """Export resources to JSON format"""
        if filename:
            try:
                clean = validate_output_filename(filename)
            except ValueError:
                raise
            if not clean.lower().endswith('.json'):
                clean = clean + '.json'
            filename = clean
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"aws_tags_scan_{timestamp}.json"

        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(resources, f, indent=2, default=str)

        self.logger.info(f"[JSON] Report exported to: {filename}")
        return filename

    def export_to_excel(self, resources: List[Dict[str, Any]], filename: str | None = None) -> str:
        """Export resources into an Excel workbook with one sheet per service type"""
        # validate/sanitize filename
        if filename:
            try:
                clean = validate_output_filename(filename)
            except ValueError:
                raise
            if not clean.lower().endswith('.xlsx'):
                clean = clean + '.xlsx'
            filename = clean
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"aws_tags_scan_{timestamp}.xlsx"

        # Group resources by service
        by_service: Dict[str, List[Dict[str, Any]]] = {}
        for r in resources:
            svc = r.get('Service', 'other')
            by_service.setdefault(svc, []).append(r)

        wb = Workbook()
        # remove default sheet
        default = wb.active
        wb.remove(default)

        for service, items in sorted(by_service.items()):
            ws = wb.create_sheet(title=service[:31])

            # Collect tag columns specific to this service
            tag_cols = self._collect_tag_columns(items)
            headers = self._build_csv_headers(tag_cols)

            # Write header
            for c_idx, h in enumerate(headers, start=1):
                ws.cell(row=1, column=c_idx, value=h)

            # Write rows
            for r_idx, resource in enumerate(items, start=2):
                row = self._build_csv_row(resource, tag_cols)
                for c_idx, val in enumerate(row, start=1):
                    # Ensure strings for Excel cells
                    if val is None:
                        cell_val = ''
                    else:
                        cell_val = str(val)
                    ws.cell(row=r_idx, column=c_idx, value=cell_val)

            # Auto-width columns (simple heuristic)
            for i, _ in enumerate(headers, start=1):
                col = get_column_letter(i)
                max_len = 0
                for cell in ws[col]:
                    if cell.value:
                        try:
                            l = len(str(cell.value))
                        except Exception:
                            l = 0
                        if l > max_len:
                            max_len = l
                ws.column_dimensions[col].width = min(max(10, max_len + 2), 60)

        wb.save(filename)
        self.logger.info(f"[XLSX] Report exported to: {filename}")
        return filename
    
    def print_summary(self, summary: Dict[str, Any]) -> None:
        """Print a formatted summary to console"""
        self.logger.info("\n" + "="*60)
        self.logger.info("AWS TAGGING SUMMARY REPORT")
        self.logger.info("="*60)

        self.logger.info("Overall Statistics:")
        self.logger.info(f"  Total Resources: {summary['total_resources']}")
        self.logger.info(f"  Resources with Tags: {summary['resources_with_tags']}")
        self.logger.info(f"  Resources without Tags: {summary['resources_without_tags']}")

        if summary['total_resources'] > 0:
            compliance_rate = (summary['resources_with_tags'] / summary['total_resources']) * 100
            self.logger.info(f"  Tag Compliance Rate: {compliance_rate:.1f}%")

        # Show policy compliance if available
        if self.tag_policy and summary['compliance']['compliant'] + summary['compliance']['non_compliant'] > 0:
            self.logger.info(f"Tag Policy Compliance ({self.tag_policy.get('name', 'Policy')}):")
            self.logger.info(f"  Compliant Resources: {summary['compliance']['compliant']}")
            self.logger.info(f"  Non-Compliant Resources: {summary['compliance']['non_compliant']}")
            self.logger.info(f"  Policy Compliance Rate: {summary['compliance']['compliance_rate']:.1f}%")

        self.logger.info("By Service:")
        for service, stats in summary['by_service'].items():
            compliance = (stats['with_tags'] / stats['total'] * 100) if stats['total'] > 0 else 0
            output = f"  {service}: {stats['total']} total, {stats['with_tags']} tagged ({compliance:.1f}%)"

            if self.tag_policy and 'compliant' in stats:
                policy_comp = (stats['compliant'] / stats['total'] * 100) if stats['total'] > 0 else 0
                output += f", {stats['compliant']} policy-compliant ({policy_comp:.1f}%)"

            self.logger.info(output)

        self.logger.info("Most Common Tags:")
        for tag, count in list(summary['common_tags'].items())[:10]:
            self.logger.info(f"  {tag}: {count} resources")

        self.logger.info("\n" + "="*60)
    
    def _print_resource_header(self, resource: Dict[str, Any]) -> None:
        """Print resource header information"""
        self.logger.info(f"[{resource['Service']} - {resource['ResourceType']}]")
        self.logger.info(f"   ID: {resource['ResourceId']}")
        self.logger.info(f"   Name: {resource['ResourceName']}")
        self.logger.info(f"   State: {resource['State']}")
        self.logger.info(f"   Tags ({resource['TagCount']}):")
    
    def _print_resource_tags(self, resource: Dict[str, Any]) -> None:
        """Print resource tags or no-tags message"""
        if resource['Tags']:
            for key, value in resource['Tags'].items():
                self.logger.info(f"     {key}: {value}")
        else:
            self.logger.info("     [No tags found]")
    
    def _print_compliance_status(self, compliance: Dict[str, Any]) -> None:
        """Print compliance status and details"""
        if compliance['compliant']:
            self.logger.info(f"   [COMPLIANT] Policy Compliant (Score: {compliance['compliance_score']:.1f}%)")
        else:
            self.logger.info(f"   [NON-COMPLIANT] Policy Non-Compliant (Score: {compliance['compliance_score']:.1f}%)")

            if compliance['missing_required']:
                self.logger.info(f"      Missing Required: {', '.join(compliance['missing_required'])}")

            if compliance['missing_recommended']:
                self.logger.info(f"      Missing Recommended: {', '.join(compliance['missing_recommended'])}")

            if compliance['invalid_values']:
                self.logger.info("      Invalid Tag Values:")
                for inv in compliance['invalid_values']:
                    self.logger.info(f"        - {inv['key']}: '{inv['value']}' (allowed: {', '.join(inv['allowed'])})")
    
    def print_detailed_report(self, resources: List[Dict[str, Any]]) -> None:
        """Print detailed resource information"""
        print("\n" + "="*80)
        print("DETAILED RESOURCE REPORT")
        print("="*80)
        
        for resource in resources:
            self._print_resource_header(resource)
            self._print_resource_tags(resource)
            
            # Show compliance status if policy is loaded
            if self.tag_policy and 'Compliance' in resource:
                self._print_compliance_status(resource['Compliance'])


def main() -> None:
    parser = argparse.ArgumentParser(description='AWS Tag Scanner - No Installation Required')
    parser.add_argument('--profile', help='AWS profile name')
    parser.add_argument('--region', help='AWS region')
    parser.add_argument('--output', choices=['console', 'csv', 'json', 'xlsx', 'all'], 
                       default='console', help='Output format')
    parser.add_argument('--detailed', action='store_true', 
                       help='Show detailed resource information')
    parser.add_argument('--filename', help='Custom output filename (without extension)')
    parser.add_argument('--dry-run', action='store_true',
                       help='Only test authentication, do not scan resources')
    parser.add_argument('--policy', help='Path to tag policy JSON file for compliance checking')
    parser.add_argument('--verbose', action='store_true', help='Enable verbose (debug) logging')
    parser.add_argument('--quiet', action='store_true', help='Quiet mode (only warnings and errors)')
    
    args = parser.parse_args()
    # Configure basic logging for the CLI tool based on CLI flags
    if args.verbose:
        log_level = logging.DEBUG
    elif args.quiet:
        log_level = logging.WARNING
    else:
        log_level = logging.INFO

    logging.basicConfig(level=log_level, format='%(levelname)s: %(message)s')
    
    try:
        scanner = AWSTagScanner(
            profile=args.profile, 
            region=args.region, 
            dry_run=args.dry_run,
            policy_file=args.policy
        )
        
        if args.dry_run:
            return
            
        resources = scanner.scan_all_resources()
        summary = scanner.generate_summary_report(resources)
        
        # Always show summary
        scanner.print_summary(summary)
        
        # Show detailed report if requested
        if args.detailed:
            scanner.print_detailed_report(resources)
        
        # Handle output options
        if args.output in ['csv', 'all']:
            scanner.export_to_csv(resources, 
                                f"{args.filename}.csv" if args.filename else None)

        if args.output in ['json', 'all']:
            scanner.export_to_json(resources, 
                                 f"{args.filename}.json" if args.filename else None)

        # If the user requested Excel output via filename extension or wants all,
        # create an .xlsx workbook with one sheet per service.
        if args.output in ['all', 'xlsx'] or (args.filename and args.filename.lower().endswith('.xlsx')):
            # prefer explicit filename without extension handling inside method
            excel_name = None
            if args.filename:
                # strip potential extension and pass base; export_to_excel will ensure .xlsx
                base = args.filename
                if base.lower().endswith('.xlsx'):
                    base = base[:-5]
                excel_name = f"{base}.xlsx"

            scanner.export_to_excel(resources, excel_name)
        
    except KeyboardInterrupt:
        print("\n[ERROR] Scan cancelled by user")
        sys.exit(1)
    except Exception as e:
        print(f"[ERROR] Unexpected error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
